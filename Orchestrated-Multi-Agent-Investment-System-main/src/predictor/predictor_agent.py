"""
Target Predictor Agent - Portfolio Validation & Performance Forecasting
Multi-Agent Investment Management System

RESPONSIBILITY: Validate investment plans and predict portfolio performance
Input: Investment plan (from Planner) + Market data (from Monitor)
Output: Validation decision (APPROVED/REVISE) with detailed feedback and reports
"""

import json
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, asdict
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import warnings
warnings.filterwarnings('ignore')

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from logger.custom_logger import CustomLogger

# Configure custom logger
logger = CustomLogger().get_logger(__file__)


@dataclass
class HistoricalAnalysis:
    """Historical backtest results"""
    predicted_return: float
    volatility: float
    sharpe_ratio: float
    max_drawdown: float
    win_rate: float
    avg_monthly_return: float
    confidence: float
    method: str = "HISTORICAL_BACKTEST"


@dataclass
class TechnicalAnalysis:
    """Current technical signal validation"""
    technical_score: float
    predicted_direction: str
    strong_signals: int
    weak_signals: int
    stock_scores: Dict[str, float]
    confidence: float
    method: str = "TECHNICAL_VALIDATION"


@dataclass
class RiskAnalysis:
    """Portfolio risk assessment"""
    risk_acceptable: bool
    volatility_ratio: float
    concentration_warning: bool
    sector_warning: bool
    risk_score: float
    warnings: List[str]
    confidence: float
    method: str = "RISK_ASSESSMENT"


@dataclass
class ValidationDecision:
    """Final validation decision"""
    customer_id: str
    validation_date: datetime
    status: str  # APPROVED or REVISE
    predicted_return: float
    target_return: float
    confidence_level: float
    feedback: List[str]
    historical_analysis: HistoricalAnalysis
    technical_analysis: TechnicalAnalysis
    risk_analysis: RiskAnalysis
    meets_expectations: bool
    revision_priority: str  # HIGH, MEDIUM, LOW


class ExcelValidationReporter:
    """Generates comprehensive Excel validation reports"""
    
    def __init__(self, reports_dir: str = "predictor_data"):
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
        self.reports_dir = os.path.join(project_root, "data", reports_dir)
        os.makedirs(self.reports_dir, exist_ok=True)
        
        # Define color scheme
        self.colors = {
            'header': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
            'approved': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'revise': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'warning': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'excellent': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
            'poor': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'neutral': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        }
        
        self.fonts = {
            'header': Font(color='FFFFFF', bold=True, size=12),
            'title': Font(bold=True, size=16),
            'subtitle': Font(bold=True, size=12),
            'normal': Font(size=11),
            'bold': Font(bold=True, size=11)
        }
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_validation_report(self, decision: ValidationDecision, 
                                 investment_plan: Dict) -> str:
        """Create comprehensive validation report"""
        
        date_str = decision.validation_date.strftime('%Y%m%d_%H%M%S')
        filename = f"{self.reports_dir}/ValidationReport_{decision.customer_id}_{date_str}.xlsx"
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create all sheets
        self._create_executive_summary(wb, decision, investment_plan)
        self._create_historical_analysis(wb, decision)
        self._create_technical_analysis(wb, decision)
        self._create_risk_assessment(wb, decision)
        self._create_recommendations(wb, decision)
        
        wb.save(filename)
        logger.info(f"Validation report saved: {filename}")
        
        return filename
    
    def _create_executive_summary(self, wb: Workbook, decision: ValidationDecision, plan: Dict):
        """Executive Summary Sheet"""
        ws = wb.create_sheet("📊 Executive Summary", 0)
        
        # Title
        ws['A1'] = "PORTFOLIO VALIDATION REPORT"
        ws['A1'].font = self.fonts['title']
        ws.merge_cells('A1:D1')
        
        # Basic Info
        ws['A3'] = "Customer ID:"
        ws['B3'] = decision.customer_id
        ws['A4'] = "Validation Date:"
        ws['B4'] = decision.validation_date.strftime('%B %d, %Y %H:%M')
        ws['A5'] = "Investment Capital:"
        ws['B5'] = f"₹{plan['total_capital']:,.2f}"
        
        # Status Box
        ws['A7'] = "VALIDATION STATUS"
        ws['A7'].font = self.fonts['subtitle']
        
        status_cell = ws['B8']
        status_cell.value = f"✅ {decision.status}" if decision.status == "APPROVED" else f"⚠️ {decision.status}"
        status_cell.font = Font(bold=True, size=14)
        status_cell.fill = self.colors['approved'] if decision.status == "APPROVED" else self.colors['revise']
        
        # Key Metrics
        ws['A10'] = "KEY METRICS"
        ws['A10'].font = self.fonts['subtitle']
        
        metrics = [
            ['Predicted Return', f"{decision.predicted_return:.2%}", decision.predicted_return >= decision.target_return],
            ['Target Return', f"{decision.target_return:.2%}", True],
            ['Confidence Level', f"{decision.confidence_level:.1%}", decision.confidence_level >= 0.85],
            ['Return Gap', f"{(decision.predicted_return - decision.target_return):.2%}", 
             decision.predicted_return >= decision.target_return]
        ]
        
        for i, (metric, value, is_good) in enumerate(metrics, 12):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'B{i}'].font = self.fonts['bold']
            
            if i > 12:  # Color code except target
                ws[f'B{i}'].fill = self.colors['approved'] if is_good else self.colors['warning']
        
        # Performance Summary
        ws['A17'] = "ANALYSIS SUMMARY"
        ws['A17'].font = self.fonts['subtitle']
        
        hist = decision.historical_analysis
        tech = decision.technical_analysis
        risk = decision.risk_analysis
        
        summary_data = [
            ['Historical Performance', f"{hist.predicted_return:.2%}", f"Sharpe: {hist.sharpe_ratio:.2f}"],
            ['Technical Signals', tech.predicted_direction, f"Score: {tech.technical_score:.2f}"],
            ['Risk Assessment', "ACCEPTABLE" if risk.risk_acceptable else "WARNING", 
             f"Score: {risk.risk_score:.2f}"]
        ]
        
        for i, (category, value, detail) in enumerate(summary_data, 19):
            ws[f'A{i}'] = category
            ws[f'B{i}'] = value
            ws[f'C{i}'] = detail
        
        # Verdict
        ws['A23'] = "VERDICT"
        ws['A23'].font = self.fonts['subtitle']
        
        if decision.status == "APPROVED":
            verdict = "✅ Portfolio meets all expectations and risk criteria. Recommended for implementation."
        else:
            verdict = f"⚠️ Portfolio requires revision. Priority: {decision.revision_priority}"
        
        ws['A24'] = verdict
        ws['A24'].font = Font(bold=True, size=11, italic=True)
        ws.merge_cells('A24:D24')
        
        self._auto_adjust_columns(ws)
    
    def _create_historical_analysis(self, wb: Workbook, decision: ValidationDecision):
        """Historical Backtest Analysis Sheet"""
        ws = wb.create_sheet("📈 Historical Analysis")
        
        ws['A1'] = "HISTORICAL BACKTEST RESULTS"
        ws['A1'].font = self.fonts['title']
        
        hist = decision.historical_analysis
        
        # Performance Metrics
        ws['A3'] = "PERFORMANCE METRICS"
        ws['A3'].font = self.fonts['subtitle']
        
        metrics_data = [
            ['Predicted Annual Return', f"{hist.predicted_return:.2%}"],
            ['Annualized Volatility', f"{hist.volatility:.2%}"],
            ['Sharpe Ratio', f"{hist.sharpe_ratio:.3f}"],
            ['Maximum Drawdown', f"{hist.max_drawdown:.2%}"],
            ['Win Rate', f"{hist.win_rate:.1%}"],
            ['Avg Monthly Return', f"{hist.avg_monthly_return:.2%}"],
            ['Confidence Level', f"{hist.confidence:.1%}"]
        ]
        
        for i, (metric, value) in enumerate(metrics_data, 5):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = self.fonts['bold']
        
        # Interpretation
        ws['A13'] = "INTERPRETATION"
        ws['A13'].font = self.fonts['subtitle']
        
        interpretations = []
        
        if hist.sharpe_ratio > 1.5:
            interpretations.append("✅ Excellent risk-adjusted returns (Sharpe > 1.5)")
        elif hist.sharpe_ratio > 1.0:
            interpretations.append("✓ Good risk-adjusted returns (Sharpe > 1.0)")
        else:
            interpretations.append("⚠️ Moderate risk-adjusted returns (Sharpe < 1.0)")
        
        if hist.max_drawdown > -0.20:
            interpretations.append("✅ Acceptable maximum drawdown (< 20%)")
        else:
            interpretations.append("⚠️ High maximum drawdown (> 20%)")
        
        if hist.win_rate > 0.60:
            interpretations.append("✅ Strong win rate (> 60%)")
        elif hist.win_rate > 0.50:
            interpretations.append("✓ Decent win rate (> 50%)")
        else:
            interpretations.append("⚠️ Below average win rate (< 50%)")
        
        for i, interp in enumerate(interpretations, 15):
            ws[f'A{i}'] = interp
        
        # Methodology
        ws['A20'] = "METHODOLOGY"
        ws['A20'].font = self.fonts['subtitle']
        
        ws['A21'] = "• Analyzed 1-year historical performance"
        ws['A22'] = "• Simulated portfolio value using actual allocations"
        ws['A23'] = "• Calculated returns, volatility, and drawdowns"
        ws['A24'] = "• Annualized metrics for comparison"
        
        self._auto_adjust_columns(ws)
    
    def _create_technical_analysis(self, wb: Workbook, decision: ValidationDecision):
        """Technical Signal Analysis Sheet"""
        ws = wb.create_sheet("🔧 Technical Analysis")
        
        ws['A1'] = "TECHNICAL SIGNAL VALIDATION"
        ws['A1'].font = self.fonts['title']
        
        tech = decision.technical_analysis
        
        # Overall Assessment
        ws['A3'] = "OVERALL TECHNICAL ASSESSMENT"
        ws['A3'].font = self.fonts['subtitle']
        
        ws['A5'] = "Technical Score:"
        ws['B5'] = f"{tech.technical_score:.2f} / 1.00"
        ws['B5'].font = self.fonts['bold']
        ws['B5'].fill = self.colors['approved'] if tech.technical_score > 0.6 else self.colors['warning']
        
        ws['A6'] = "Predicted Direction:"
        ws['B6'] = tech.predicted_direction
        ws['B6'].font = self.fonts['bold']
        ws['B6'].fill = self.colors['approved'] if tech.predicted_direction == "POSITIVE" else self.colors['revise']
        
        ws['A7'] = "Strong Signals:"
        ws['B7'] = tech.strong_signals
        
        ws['A8'] = "Weak Signals:"
        ws['B8'] = tech.weak_signals
        
        ws['A9'] = "Confidence Level:"
        ws['B9'] = f"{tech.confidence:.1%}"
        
        # Stock-by-Stock Scores
        ws['A11'] = "STOCK-BY-STOCK TECHNICAL SCORES"
        ws['A11'].font = self.fonts['subtitle']
        
        headers = ['Symbol', 'Technical Score', 'Rating']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=13, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        sorted_stocks = sorted(tech.stock_scores.items(), key=lambda x: x[1], reverse=True)
        
        for i, (symbol, score) in enumerate(sorted_stocks, 14):
            ws[f'A{i}'] = symbol
            ws[f'B{i}'] = f"{score:.3f}"
            
            if score > 0.7:
                rating = "STRONG"
                fill = self.colors['excellent']
            elif score > 0.5:
                rating = "GOOD"
                fill = self.colors['approved']
            elif score > 0.3:
                rating = "MODERATE"
                fill = self.colors['warning']
            else:
                rating = "WEAK"
                fill = self.colors['revise']
            
            ws[f'C{i}'] = rating
            ws[f'C{i}'].fill = fill
        
        # Signal Interpretation
        ws['A25'] = "SIGNAL INTERPRETATION"
        ws['A25'].font = self.fonts['subtitle']
        
        ws['A26'] = "• Technical score > 0.7: Strong buy signals"
        ws['A27'] = "• Technical score 0.5-0.7: Moderate buy signals"
        ws['A28'] = "• Technical score 0.3-0.5: Neutral signals"
        ws['A29'] = "• Technical score < 0.3: Weak/sell signals"
        
        self._auto_adjust_columns(ws)
    
    def _create_risk_assessment(self, wb: Workbook, decision: ValidationDecision):
        """Risk Assessment Sheet"""
        ws = wb.create_sheet("⚠️ Risk Assessment")
        
        ws['A1'] = "PORTFOLIO RISK ANALYSIS"
        ws['A1'].font = self.fonts['title']
        
        risk = decision.risk_analysis
        
        # Overall Risk Status
        ws['A3'] = "OVERALL RISK STATUS"
        ws['A3'].font = self.fonts['subtitle']
        
        status_cell = ws['B4']
        status_cell.value = "✅ ACCEPTABLE" if risk.risk_acceptable else "⚠️ WARNING"
        status_cell.font = Font(bold=True, size=12)
        status_cell.fill = self.colors['approved'] if risk.risk_acceptable else self.colors['revise']
        
        # Risk Metrics
        ws['A6'] = "RISK METRICS"
        ws['A6'].font = self.fonts['subtitle']
        
        metrics = [
            ['Risk Score', f"{risk.risk_score:.2f} / 1.00", risk.risk_score > 0.7],
            ['Volatility Ratio', f"{risk.volatility_ratio:.2f}x", risk.volatility_ratio <= 1.0],
            ['Concentration Risk', "WARNING" if risk.concentration_warning else "OK", 
             not risk.concentration_warning],
            ['Sector Concentration', "WARNING" if risk.sector_warning else "OK", 
             not risk.sector_warning]
        ]
        
        for i, (metric, value, is_good) in enumerate(metrics, 8):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = self.fonts['bold']
            ws[f'B{i}'].fill = self.colors['approved'] if is_good else self.colors['warning']
        
        # Risk Warnings
        ws['A13'] = "RISK WARNINGS"
        ws['A13'].font = self.fonts['subtitle']
        
        if risk.warnings:
            for i, warning in enumerate(risk.warnings, 15):
                ws[f'A{i}'] = f"⚠️ {warning}"
                ws[f'A{i}'].fill = self.colors['warning']
        else:
            ws['A15'] = "✅ No risk warnings detected"
            ws['A15'].fill = self.colors['approved']
        
        # Risk Guidelines
        ws['A20'] = "RISK GUIDELINES"
        ws['A20'].font = self.fonts['subtitle']
        
        ws['A21'] = "• Volatility should match customer risk tolerance"
        ws['A22'] = "• No single stock should exceed 15-20% allocation"
        ws['A23'] = "• Sector concentration should not exceed 40%"
        ws['A24'] = "• Portfolio should maintain proper diversification"
        
        self._auto_adjust_columns(ws)
    
    def _create_recommendations(self, wb: Workbook, decision: ValidationDecision):
        """Recommendations Sheet"""
        ws = wb.create_sheet("💡 Recommendations")
        
        ws['A1'] = "VALIDATION FEEDBACK & RECOMMENDATIONS"
        ws['A1'].font = self.fonts['title']
        
        # Status
        ws['A3'] = "Decision:"
        ws['B3'] = decision.status
        ws['B3'].font = Font(bold=True, size=12)
        ws['B3'].fill = self.colors['approved'] if decision.status == "APPROVED" else self.colors['revise']
        
        if decision.status == "REVISE":
            ws['A4'] = "Priority:"
            ws['B4'] = decision.revision_priority
            ws['B4'].font = Font(bold=True)
        
        # Feedback
        ws['A6'] = "DETAILED FEEDBACK"
        ws['A6'].font = self.fonts['subtitle']
        
        for i, feedback_item in enumerate(decision.feedback, 8):
            ws[f'A{i}'] = feedback_item
            ws.merge_cells(f'A{i}:D{i}')
            ws[f'A{i}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Next Steps
        next_row = 8 + len(decision.feedback) + 2
        ws[f'A{next_row}'] = "NEXT STEPS"
        ws[f'A{next_row}'].font = self.fonts['subtitle']
        
        if decision.status == "APPROVED":
            next_steps = [
                "✅ Proceed with portfolio implementation",
                "✅ Monitor performance against predictions",
                "✅ Review quarterly for rebalancing needs",
                "✅ Maintain risk management discipline"
            ]
        else:
            next_steps = [
                "⚠️ Review feedback points carefully",
                "⚠️ Adjust portfolio allocation as suggested",
                "⚠️ Re-run planner with updated parameters",
                "⚠️ Resubmit for validation"
            ]
        
        for i, step in enumerate(next_steps):
            ws[f'A{next_row + 2 + i}'] = step
        
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width


class TargetPredictorAgent:
    """Target Predictor Agent - Portfolio Validation Engine"""
    
    CONFIDENCE_THRESHOLD = 0.85  # 85% confidence requirement
    
    def __init__(self):
        self.logger = logger
        self.excel_reporter = ExcelValidationReporter()
        self.logger.info("Target Predictor Agent initialized")
    
    def load_investment_plan(self, customer_id: str) -> Dict:
        """Load investment plan from Planner"""
        try:
            current_file = os.path.abspath(__file__)
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
            plan_path = os.path.join(project_root, "data", "planner_data", 
                                    f"investment_plan_{customer_id}.json")
            
            with open(plan_path, 'r') as f:
                plan = json.load(f)
            
            self.logger.info(f"Loaded investment plan for {customer_id}")
            return plan
            
        except Exception as e:
            self.logger.error(f"Failed to load investment plan: {e}")
            raise
    
    def load_current_market_data(self) -> Dict:
        """Load current market intelligence from Monitor"""
        try:
            current_file = os.path.abspath(__file__)
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
            market_path = os.path.join(project_root, "data", "monitor_data", 
                                      "general_market_data.json")
            
            with open(market_path, 'r') as f:
                market_data = json.load(f)
            
            self.logger.info("Loaded current market data from Monitor")
            return market_data
            
        except Exception as e:
            self.logger.error(f"Failed to load market data: {e}")
            raise
    
    def fetch_historical_prices(self, symbols: List[str], period: str = "1y") -> Dict[str, pd.Series]:
        """Fetch historical prices for backtesting"""
        self.logger.info(f"Fetching historical data for {len(symbols)} stocks...")
        
        historical_data = {}
        
        for symbol in symbols:
            try:
                ticker = yf.Ticker(symbol)
                hist = ticker.history(period=period)
                
                if not hist.empty:
                    historical_data[symbol] = hist['Close']
                    self.logger.info(f"✓ {symbol}: {len(hist)} days")
                else:
                    self.logger.warning(f"✗ {symbol}: No data")
                    
            except Exception as e:
                self.logger.error(f"✗ {symbol}: {e}")
                continue
        
        return historical_data
    
    def backtest_historical_performance(self, plan: Dict, 
                                       historical_data: Dict[str, pd.Series]) -> HistoricalAnalysis:
        """Backtest portfolio performance using historical data"""
        
        self.logger.info("Running historical backtest...")
        
        # Get all dates (intersection of all stocks)
        all_dates = None
        for symbol in historical_data.keys():
            if all_dates is None:
                all_dates = historical_data[symbol].index
            else:
                all_dates = all_dates.intersection(historical_data[symbol].index)
        
        if len(all_dates) < 30:
            self.logger.warning("Insufficient historical data")
            return HistoricalAnalysis(
                predicted_return=0.0,
                volatility=0.0,
                sharpe_ratio=0.0,
                max_drawdown=0.0,
                win_rate=0.0,
                avg_monthly_return=0.0,
                confidence=0.3
            )
        
        # Calculate portfolio value over time
        portfolio_values = []
        
        for date in all_dates:
            daily_value = 0
            for allocation in plan['allocations']:
                symbol = allocation['symbol']
                if symbol in historical_data:
                    price = historical_data[symbol][date]
                    shares = allocation['shares']
                    daily_value += shares * price
            
            portfolio_values.append(daily_value)
        
        portfolio_values = np.array(portfolio_values)
        
        # Calculate returns
        returns = np.diff(portfolio_values) / portfolio_values[:-1]
        
        # Metrics
        with np.errstate(divide='ignore', invalid='ignore'):
            annual_return = np.mean(returns) * 252
            volatility = np.std(returns) * np.sqrt(252)
            sharpe_ratio = annual_return / volatility if volatility > 0 else 0.0
        
        # Handle NaN/Inf
        if not np.isfinite(annual_return): annual_return = 0.0
        if not np.isfinite(volatility): volatility = 0.0
        if not np.isfinite(sharpe_ratio): sharpe_ratio = 0.0
        
        # Max drawdown
        with np.errstate(divide='ignore', invalid='ignore'):
            cumulative = (1 + returns).cumprod()
            running_max = np.maximum.accumulate(cumulative)
            drawdown = (cumulative - running_max) / running_max
            max_drawdown = np.min(drawdown)
            
        if not np.isfinite(max_drawdown): max_drawdown = 0.0
        
        # Win rate (% of positive return days)
        win_rate = np.sum(returns > 0) / len(returns)
        
        # Average monthly return
        monthly_returns = []
        current_month_returns = []
        current_month = all_dates[0].month
        
        for i, date in enumerate(all_dates[1:]):
            if date.month != current_month:
                if current_month_returns:
                    monthly_return = np.prod([1 + r for r in current_month_returns]) - 1
                    monthly_returns.append(monthly_return)
                current_month_returns = []
                current_month = date.month
            current_month_returns.append(returns[i])
        
        avg_monthly_return = np.mean(monthly_returns) if monthly_returns else 0.0
        if not np.isfinite(avg_monthly_return): avg_monthly_return = 0.0
        
        # Confidence based on data quality
        confidence = min(0.75, 0.5 + (len(all_dates) / 365) * 0.25)
        
        self.logger.info(f"Backtest complete: Return={annual_return:.2%}, Sharpe={sharpe_ratio:.2f}")
        
        return HistoricalAnalysis(
            predicted_return=annual_return,
            volatility=volatility,
            sharpe_ratio=sharpe_ratio,
            max_drawdown=max_drawdown,
            win_rate=win_rate,
            avg_monthly_return=avg_monthly_return,
            confidence=confidence
        )
    
    def validate_current_signals(self, plan: Dict, 
                                market_data: Dict) -> TechnicalAnalysis:
        """Validate portfolio using Monitor's current technical analysis"""
        
        self.logger.info("Validating current technical signals...")
        
        stock_scores = {}
        strong_signals = 0
        weak_signals = 0
        
        for allocation in plan['allocations']:
            symbol = allocation['symbol']
            weight = allocation['weight']
            
            # Find stock in market data
            stock_data = next(
                (s for s in market_data['stocks'] if s['symbol'] == symbol),
                None
            )
            
            if not stock_data:
                stock_scores[symbol] = 0.5  # Neutral if no data
                continue
            
            tech = stock_data.get('technical_indicators', {})
            
            # Calculate score based on technical indicators
            score = 0.5  # Start neutral
            
            # RSI scoring
            rsi = tech.get('rsi', 50)
            if 40 <= rsi <= 60:
                score += 0.15  # Healthy range
            elif 30 <= rsi <= 70:
                score += 0.10  # Acceptable range
            elif rsi > 80 or rsi < 20:
                score -= 0.15  # Extreme (overbought/oversold)
            
            # Momentum scoring
            momentum = tech.get('momentum_20d', 0)
            if momentum > 0.05:
                score += 0.20
            elif momentum > 0:
                score += 0.10
            elif momentum < -0.05:
                score -= 0.20
            elif momentum < 0:
                score -= 0.10
            
            # Signal strength scoring
            signal = tech.get('signal_strength', 'NEUTRAL')
            if signal == 'STRONG_UP':
                score += 0.20
                strong_signals += 1
            elif signal == 'UP':
                score += 0.10
            elif signal == 'DOWN':
                score -= 0.10
                weak_signals += 1
            elif signal == 'STRONG_DOWN':
                score -= 0.20
                weak_signals += 1
            
            # Normalize to 0-1 range
            score = max(0, min(1, score))
            stock_scores[symbol] = score
        
        # Calculate weighted portfolio score
        total_score = sum(
            stock_scores[alloc['symbol']] * alloc['weight']
            for alloc in plan['allocations']
            if alloc['symbol'] in stock_scores
        )
        
        predicted_direction = "POSITIVE" if total_score > 0.5 else "NEGATIVE"
        
        self.logger.info(f"Technical validation: Score={total_score:.2f}, Direction={predicted_direction}")
        
        return TechnicalAnalysis(
            technical_score=total_score,
            predicted_direction=predicted_direction,
            strong_signals=strong_signals,
            weak_signals=weak_signals,
            stock_scores=stock_scores,
            confidence=0.70
        )
    
    def assess_portfolio_risk(self, plan: Dict) -> RiskAnalysis:
        """Assess portfolio risk against customer tolerance"""
        
        self.logger.info("Assessing portfolio risk...")
        
        risk_analysis = plan['risk_analysis']
        portfolio_vol = risk_analysis['portfolio_volatility']
        concentration_risk = risk_analysis['concentration_risk']
        sector_dist = risk_analysis['sector_distribution']
        
        # Determine customer risk tolerance from plan
        risk_level = risk_analysis['risk_level']
        
        # Risk thresholds
        risk_thresholds = {
            'LOW': 0.20,
            'MEDIUM': 0.30,
            'HIGH': 0.50
        }
        
        max_acceptable_vol = risk_thresholds.get(risk_level, 0.30)
        
        # Check various risk factors
        warnings = []
        risk_score = 1.0
        
        # Volatility check
        volatility_ratio = portfolio_vol / max_acceptable_vol
        if portfolio_vol > max_acceptable_vol:
            warnings.append(f"Portfolio volatility ({portfolio_vol:.2%}) exceeds {risk_level} risk tolerance ({max_acceptable_vol:.2%})")
            risk_score -= 0.3
        
        # Concentration check
        concentration_warning = concentration_risk > 0.20
        if concentration_warning:
            warnings.append(f"High concentration risk: {concentration_risk:.2%} in single stock (recommended < 20%)")
            risk_score -= 0.2
        
        # Sector concentration check
        sector_warning = any(w > 0.40 for w in sector_dist.values())
        if sector_warning:
            max_sector = max(sector_dist.values())
            sector_name = [k for k, v in sector_dist.items() if v == max_sector][0]
            warnings.append(f"High sector concentration: {max_sector:.2%} in {sector_name} (recommended < 40%)")
            risk_score -= 0.2
        
        # Diversification check
        num_stocks = plan['portfolio_metrics']['number_of_stocks']
        if num_stocks < 5:
            warnings.append(f"Low diversification: only {num_stocks} stocks (recommended >= 5)")
            risk_score -= 0.2
        
        risk_score = max(0, risk_score)
        risk_acceptable = risk_score >= 0.6 and portfolio_vol <= max_acceptable_vol
        
        self.logger.info(f"Risk assessment: Score={risk_score:.2f}, Acceptable={risk_acceptable}")
        
        return RiskAnalysis(
            risk_acceptable=risk_acceptable,
            volatility_ratio=volatility_ratio,
            concentration_warning=concentration_warning,
            sector_warning=sector_warning,
            risk_score=risk_score,
            warnings=warnings,
            confidence=0.90
        )
    
    def make_validation_decision(self, 
                                historical: HistoricalAnalysis,
                                technical: TechnicalAnalysis,
                                risk: RiskAnalysis,
                                plan: Dict,
                                customer_id: str) -> ValidationDecision:
        """Make final validation decision with ensemble approach"""
        
        self.logger.info("Making validation decision...")
        
        # Extract target return
        target_return = plan.get('expected_annual_return', 0)
        
        # Weighted ensemble prediction
        # Historical: 60%, Technical: 30%, Risk: 10%
        ensemble_return = (
            historical.predicted_return * 0.60 +
            technical.technical_score * 0.15 * 0.30  # Scale technical score to return range
        )
        
        # Weighted confidence
        ensemble_confidence = (
            historical.confidence * 0.60 +
            technical.confidence * 0.30 +
            risk.confidence * 0.10
        )
        
        # Decision criteria (ALL must be true for APPROVED)
        meets_return = ensemble_return >= target_return
        meets_confidence = ensemble_confidence >= self.CONFIDENCE_THRESHOLD
        meets_risk = risk.risk_acceptable
        
        meets_expectations = meets_return and meets_confidence and meets_risk
        status = "APPROVED" if meets_expectations else "REVISE"
        
        # Determine revision priority
        if status == "REVISE":
            if not meets_risk:
                revision_priority = "HIGH"
            elif ensemble_return < target_return * 0.8:
                revision_priority = "HIGH"
            elif not meets_confidence:
                revision_priority = "MEDIUM"
            else:
                revision_priority = "LOW"
        else:
            revision_priority = "NONE"
        
        # Generate detailed feedback
        feedback = self._generate_feedback(
            status, ensemble_return, target_return,
            historical, technical, risk,
            meets_return, meets_confidence, meets_risk
        )
        
        self.logger.info(f"Decision: {status} | Return: {ensemble_return:.2%} | Confidence: {ensemble_confidence:.1%}")
        
        return ValidationDecision(
            customer_id=customer_id,
            validation_date=datetime.now(),
            status=status,
            predicted_return=ensemble_return,
            target_return=target_return,
            confidence_level=ensemble_confidence,
            feedback=feedback,
            historical_analysis=historical,
            technical_analysis=technical,
            risk_analysis=risk,
            meets_expectations=meets_expectations,
            revision_priority=revision_priority
        )
    
    def _generate_feedback(self, status: str, predicted_return: float, 
                          target_return: float,
                          historical: HistoricalAnalysis,
                          technical: TechnicalAnalysis,
                          risk: RiskAnalysis,
                          meets_return: bool,
                          meets_confidence: bool,
                          meets_risk: bool) -> List[str]:
        """Generate detailed feedback in bullet points"""
        
        feedback = []
        
        if status == "APPROVED":
            # Positive feedback
            feedback.append(f"✅ Portfolio meets return expectations: Predicted {predicted_return:.2%} vs Target {target_return:.2%}")
            feedback.append(f"✅ High confidence level: {historical.confidence * 0.6 + technical.confidence * 0.3 + risk.confidence * 0.1:.1%} (threshold: 85%)")
            feedback.append(f"✅ Risk profile acceptable: Portfolio volatility {historical.volatility:.2%}")
            
            if historical.sharpe_ratio > 1.5:
                feedback.append(f"✅ Excellent risk-adjusted returns: Sharpe ratio {historical.sharpe_ratio:.2f}")
            
            if technical.technical_score > 0.7:
                feedback.append(f"✅ Strong technical signals: Score {technical.technical_score:.2f}")
            
            feedback.append("✅ Recommendation: Proceed with portfolio implementation")
            
        else:
            # Issues to address
            feedback.append(f"⚠️ Portfolio requires revision - Priority: {risk.risk_score}")
            
            # Return gap
            if not meets_return:
                gap = target_return - predicted_return
                feedback.append(f"❌ Return shortfall: {gap:.2%} below target")
                feedback.append(f"   → Predicted return: {predicted_return:.2%} | Target: {target_return:.2%}")
                
                # Specific suggestions
                if historical.predicted_return < target_return:
                    feedback.append(f"   → Historical performance below target: Consider selecting higher-performing stocks")
                
                if technical.technical_score < 0.5:
                    feedback.append(f"   → Weak technical signals detected: Replace stocks with better momentum")
                    feedback.append(f"   → Stocks with weak signals: {technical.weak_signals}/{technical.strong_signals + technical.weak_signals}")
            
            # Confidence issues
            if not meets_confidence:
                feedback.append(f"❌ Confidence level below threshold: {(historical.confidence * 0.6 + technical.confidence * 0.3 + risk.confidence * 0.1):.1%} < 85%")
                
                if historical.confidence < 0.7:
                    feedback.append(f"   → Limited historical data reliability")
                
                if technical.predicted_direction == "NEGATIVE":
                    feedback.append(f"   → Technical indicators showing negative momentum")
            
            # Risk issues
            if not meets_risk:
                feedback.append(f"❌ Risk profile exceeds acceptable limits")
                
                for warning in risk.warnings:
                    feedback.append(f"   → {warning}")
                
                if risk.volatility_ratio > 1.2:
                    feedback.append(f"   → Action: Reduce allocation to high-volatility stocks")
                
                if risk.concentration_warning:
                    feedback.append(f"   → Action: Distribute concentration across more stocks")
                
                if risk.sector_warning:
                    feedback.append(f"   → Action: Reduce sector concentration through diversification")
            
            # Performance issues
            if historical.sharpe_ratio < 1.0:
                feedback.append(f"⚠️ Moderate risk-adjusted returns: Sharpe ratio {historical.sharpe_ratio:.2f}")
                feedback.append(f"   → Action: Seek stocks with better risk-return profile")
            
            if historical.max_drawdown < -0.20:
                feedback.append(f"⚠️ High maximum drawdown: {historical.max_drawdown:.2%}")
                feedback.append(f"   → Action: Include more defensive stocks to reduce drawdown risk")
            
            # Technical signal issues
            if technical.weak_signals > technical.strong_signals:
                feedback.append(f"⚠️ More weak signals ({technical.weak_signals}) than strong signals ({technical.strong_signals})")
                feedback.append(f"   → Action: Replace stocks with DOWN/STRONG_DOWN signals")
            
            # Specific stock recommendations
            weak_stocks = [symbol for symbol, score in technical.stock_scores.items() if score < 0.4]
            if weak_stocks:
                feedback.append(f"⚠️ Stocks with weak technical signals: {', '.join(weak_stocks[:3])}")
                feedback.append(f"   → Action: Consider replacing with stocks showing stronger momentum")
            
            feedback.append("⚠️ Recommendation: Revise portfolio allocation and resubmit for validation")
        
        return feedback
    
    def validate_investment_plan(self, customer_id: str) -> ValidationDecision:
        """Main validation pipeline"""
        
        self.logger.info(f"Starting validation for customer: {customer_id}")
        
        try:
            # 1. Load inputs
            plan = self.load_investment_plan(customer_id)
            current_market = self.load_current_market_data()
            
            # 2. Fetch historical data
            symbols = [alloc['symbol'] for alloc in plan['allocations']]
            historical_data = self.fetch_historical_prices(symbols, period="1y")
            
            # 3. Run analyses
            historical = self.backtest_historical_performance(plan, historical_data)
            technical = self.validate_current_signals(plan, current_market)
            risk = self.assess_portfolio_risk(plan)
            
            # 4. Make decision
            decision = self.make_validation_decision(
                historical, technical, risk, plan, customer_id
            )
            
            # 5. Export results
            self.export_validation_json(decision, customer_id)
            report_file = self.excel_reporter.create_validation_report(decision, plan)
            
            # 6. Print summary
            self.print_validation_summary(decision, report_file)
            
            self.logger.info("Validation completed successfully")
            return decision
            
        except Exception as e:
            self.logger.error(f"Validation failed: {e}")
            raise
    
    def export_validation_json(self, decision: ValidationDecision, customer_id: str):
        """Export validation decision to JSON for downstream agents"""
        
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
        export_dir = os.path.join(project_root, "data", "predictor_data")
        os.makedirs(export_dir, exist_ok=True)
        
        validation_data = {
            "customer_id": decision.customer_id,
            "validation_date": decision.validation_date.isoformat(),
            "status": decision.status,
            "predicted_return": float(decision.predicted_return),
            "target_return": float(decision.target_return),
            "confidence_level": float(decision.confidence_level),
            "meets_expectations": bool(decision.meets_expectations),
            "revision_priority": decision.revision_priority,
            "feedback": decision.feedback,
            "historical_analysis": {
                "predicted_return": float(decision.historical_analysis.predicted_return),
                "volatility": float(decision.historical_analysis.volatility),
                "sharpe_ratio": float(decision.historical_analysis.sharpe_ratio),
                "max_drawdown": float(decision.historical_analysis.max_drawdown),
                "win_rate": float(decision.historical_analysis.win_rate),
                "confidence": float(decision.historical_analysis.confidence)
            },
            "technical_analysis": {
                "technical_score": float(decision.technical_analysis.technical_score),
                "predicted_direction": decision.technical_analysis.predicted_direction,
                "strong_signals": int(decision.technical_analysis.strong_signals),
                "weak_signals": int(decision.technical_analysis.weak_signals),
                "confidence": float(decision.technical_analysis.confidence)
            },
            "risk_analysis": {
                "risk_acceptable": bool(decision.risk_analysis.risk_acceptable),
                "volatility_ratio": float(decision.risk_analysis.volatility_ratio),
                "concentration_warning": bool(decision.risk_analysis.concentration_warning),
                "sector_warning": bool(decision.risk_analysis.sector_warning),
                "risk_score": float(decision.risk_analysis.risk_score),
                "warnings": list(decision.risk_analysis.warnings),
                "confidence": float(decision.risk_analysis.confidence)
            }
        }
        
        filename = f"{export_dir}/validation_{customer_id}.json"
        with open(filename, 'w') as f:
            json.dump(validation_data, f, indent=2)
        
        self.logger.info(f"Validation JSON exported: {filename}")
    
    def print_validation_summary(self, decision: ValidationDecision, report_file: str):
        """Print readable validation summary"""
        
        print("\n" + "="*80)
        print(f"🎯 PORTFOLIO VALIDATION SUMMARY - {decision.customer_id}")
        print("="*80)
        
        # Status
        status_symbol = "✅" if decision.status == "APPROVED" else "⚠️"
        print(f"\n{status_symbol} VALIDATION STATUS: {decision.status}")
        
        if decision.status == "REVISE":
            print(f"   Priority: {decision.revision_priority}")
        
        # Key Metrics
        print(f"\n📊 KEY METRICS:")
        print(f"   Predicted Return:  {decision.predicted_return:.2%}")
        print(f"   Target Return:     {decision.target_return:.2%}")
        print(f"   Return Gap:        {(decision.predicted_return - decision.target_return):.2%}")
        print(f"   Confidence Level:  {decision.confidence_level:.1%} (threshold: 85%)")
        
        # Analysis Results
        print(f"\n📈 HISTORICAL ANALYSIS:")
        hist = decision.historical_analysis
        print(f"   Annual Return:     {hist.predicted_return:.2%}")
        print(f"   Volatility:        {hist.volatility:.2%}")
        print(f"   Sharpe Ratio:      {hist.sharpe_ratio:.3f}")
        print(f"   Max Drawdown:      {hist.max_drawdown:.2%}")
        print(f"   Win Rate:          {hist.win_rate:.1%}")
        
        print(f"\n🔧 TECHNICAL ANALYSIS:")
        tech = decision.technical_analysis
        print(f"   Technical Score:   {tech.technical_score:.2f} / 1.00")
        print(f"   Direction:         {tech.predicted_direction}")
        print(f"   Strong Signals:    {tech.strong_signals}")
        print(f"   Weak Signals:      {tech.weak_signals}")
        
        print(f"\n⚠️ RISK ASSESSMENT:")
        risk = decision.risk_analysis
        print(f"   Risk Status:       {'ACCEPTABLE' if risk.risk_acceptable else 'WARNING'}")
        print(f"   Risk Score:        {risk.risk_score:.2f} / 1.00")
        print(f"   Volatility Ratio:  {risk.volatility_ratio:.2f}x")
        
        if risk.warnings:
            print(f"   Warnings:")
            for warning in risk.warnings:
                print(f"      • {warning}")
        
        # Feedback
        print(f"\n💡 DETAILED FEEDBACK:")
        for feedback_item in decision.feedback:
            print(f"   {feedback_item}")
        
        # Report location
        print(f"\n📄 REPORTS GENERATED:")
        print(f"   Excel Report: {report_file}")
        print(f"   JSON Output:  data/predictor_data/validation_{decision.customer_id}.json")
        
        print("\n" + "="*80)


def main():
    """Demo: Validate investment plan for sample customer"""
    
    logger.info("Starting Target Predictor Agent Demo")
    
    # Customer ID to validate
    customer_id = "CUST_001"
    
    # Create predictor
    predictor = TargetPredictorAgent()
    
    try:
        # Run validation
        decision = predictor.validate_investment_plan(customer_id)
        
        print("\n" + "="*80)
        if decision.status == "APPROVED":
            print("🎉 VALIDATION SUCCESSFUL!")
            print(f"✅ Portfolio approved with {decision.confidence_level:.1%} confidence")
            print(f"✅ Predicted return: {decision.predicted_return:.2%}")
        else:
            print("⚠️ VALIDATION REQUIRES REVISION")
            print(f"⚠️ Priority: {decision.revision_priority}")
            print(f"⚠️ Please review feedback and adjust portfolio")
        print("="*80)
        
    except Exception as e:
        logger.error(f"Error in validation: {e}")
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    main()
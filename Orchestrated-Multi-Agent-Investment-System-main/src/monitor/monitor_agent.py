"""
Monitor Agent - Pure Market Data Collection (Refactored)
Multi-Agent Investment Management System

RESPONSIBILITY: Collect and analyze market data ONLY
NO customer-specific logic - purely market intelligence
"""

import yfinance as yf
import pandas as pd
import numpy as np
import json
from datetime import datetime
from typing import List
from dataclasses import dataclass, asdict
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import warnings
warnings.filterwarnings('ignore')

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from logger.custom_logger import CustomLogger

# Configure custom logger
logger = CustomLogger().get_logger(__file__)

@dataclass
class StockData:
    """Pure stock data"""
    symbol: str
    company_name: str
    sector: str
    current_price: float
    price_change: float
    price_change_pct: float
    volume: int
    market_cap: float
    pe_ratio: float
    open_price: float
    high_price: float
    low_price: float
    date: str

@dataclass
class TechnicalIndicators:
    """Technical analysis indicators"""
    symbol: str
    rsi: float
    sma_20: float
    sma_50: float
    sma_200: float
    volatility: float
    momentum_20d: float
    momentum_50d: float
    bollinger_upper: float
    bollinger_lower: float
    macd: float
    signal_strength: str

@dataclass
class SectorAnalysis:
    """Sector performance analysis"""
    sector: str
    stock_count: int
    avg_price_change: float
    avg_volume_change: float
    sector_volatility: float
    trend_direction: str
    top_performers: List[str]
    bottom_performers: List[str]
    sector_strength: str

@dataclass
class MarketOverview:
    """Overall market condition assessment"""
    analysis_date: datetime
    market_sentiment: str
    market_volatility: float
    advancing_stocks: int
    declining_stocks: int
    total_volume: int
    market_breadth: float
    fear_greed_index: str


class ExcelMarketReporter:
    """Handles Excel report generation for general market data ONLY"""
    
    def __init__(self, reports_dir: str = "monitor_data"):
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
        self.reports_dir = os.path.join(project_root, "data", reports_dir)
        os.makedirs(self.reports_dir, exist_ok=True)
        
        self.colors = {
            'header': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
            'positive': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'negative': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'neutral': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'strong_up': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
            'strong_down': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'sector_tech': PatternFill(start_color='D4E6F1', end_color='D4E6F1', fill_type='solid'),
            'sector_banking': PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid'),
            'sector_pharma': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
            'sector_auto': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'sector_energy': PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid'),
            'sector_fmcg': PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
        }
        
        self.fonts = {
            'header': Font(color='FFFFFF', bold=True, size=12),
            'title': Font(bold=True, size=14),
            'subtitle': Font(bold=True, size=11),
        }
    
    def create_market_report(self, market_overview: MarketOverview, 
                           stock_data: List[StockData],
                           technical_indicators: List[TechnicalIndicators],
                           sector_analysis: List[SectorAnalysis]) -> str:
        """Create comprehensive market report - NO customer data"""
        
        date_str = market_overview.analysis_date.strftime('%Y%m%d')
        filename = f"{self.reports_dir}/MarketData_General_{date_str}.xlsx"
        
        wb = Workbook()
        wb.remove(wb.active)
        
        self._create_market_overview_sheet(wb, market_overview)
        self._create_stock_data_sheet(wb, stock_data)
        self._create_technical_indicators_sheet(wb, technical_indicators)
        self._create_sector_analysis_sheet(wb, sector_analysis)
        self._create_historical_trends_sheet(wb, stock_data, sector_analysis)
        
        wb.save(filename)
        logger.info(f"Market report saved: {filename}")
        
        return filename
    
    def _create_market_overview_sheet(self, wb: Workbook, market_overview: MarketOverview):
        ws = wb.create_sheet("📊 Market Overview", 0)
        
        ws['A1'] = "GENERAL MARKET DATA ANALYSIS"
        ws['A2'] = f"Analysis Date: {market_overview.analysis_date.strftime('%B %d, %Y %H:%M')}"
        ws['A3'] = "Comprehensive Market Analysis - All Sectors"
        
        for row in range(1, 4):
            ws[f'A{row}'].font = self.fonts['title']
        
        ws['A5'] = "📈 MARKET SENTIMENT"
        ws['A5'].font = self.fonts['subtitle']
        
        sentiment_color = (self.colors['positive'] if market_overview.market_sentiment == 'BULLISH' 
                          else self.colors['negative'] if market_overview.market_sentiment == 'BEARISH' 
                          else self.colors['neutral'])
        
        sentiment_data = [
            ['Overall Market Sentiment', market_overview.market_sentiment],
            ['Market Volatility', f"{market_overview.market_volatility:.2%}"],
            ['Fear/Greed Index', market_overview.fear_greed_index],
            ['Market Breadth Ratio', f"{market_overview.market_breadth:.2f}"]
        ]
        
        for i, (metric, value) in enumerate(sentiment_data, 7):
            ws[f'A{i}'] = metric
            cell = ws[f'B{i}']
            cell.value = value
            if i == 7:
                cell.fill = sentiment_color
                cell.font = Font(bold=True)
        
        ws['A13'] = "📊 MARKET STATISTICS"
        ws['A13'].font = self.fonts['subtitle']
        
        stats_data = [
            ['Advancing Stocks', market_overview.advancing_stocks],
            ['Declining Stocks', market_overview.declining_stocks],
            ['Total Trading Volume', f"{market_overview.total_volume:,}"],
            ['Net Advancers', market_overview.advancing_stocks - market_overview.declining_stocks]
        ]
        
        for i, (metric, value) in enumerate(stats_data, 15):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
        
        self._auto_adjust_columns(ws)
    
    def _create_stock_data_sheet(self, wb: Workbook, stock_data: List[StockData]):
        ws = wb.create_sheet("💰 Stock Data")
        
        ws['A1'] = "COMPREHENSIVE STOCK DATA"
        ws['A1'].font = self.fonts['title']
        
        if not stock_data:
            ws['A3'] = "No stock data available"
            return
        
        df = pd.DataFrame([{
            'Symbol': s.symbol,
            'Company Name': s.company_name,
            'Sector': s.sector,
            'Current Price (₹)': s.current_price,
            'Open (₹)': s.open_price,
            'High (₹)': s.high_price,
            'Low (₹)': s.low_price,
            'Volume': s.volume,
            'Day Change (₹)': s.price_change,
            'Day Change %': s.price_change_pct,
            'Market Cap (₹Cr)': s.market_cap / 10000000 if s.market_cap else 0,
            'P/E Ratio': s.pe_ratio,
            'Date': s.date
        } for s in stock_data])
        
        df = df.sort_values(['Sector', 'Day Change %'], ascending=[True, False])
        
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        sector_colors = {
            'TECH': self.colors['sector_tech'],
            'BANKING': self.colors['sector_banking'], 
            'PHARMA': self.colors['sector_pharma'],
            'AUTO': self.colors['sector_auto'],
            'ENERGY': self.colors['sector_energy'],
            'FMCG': self.colors['sector_fmcg']
        }
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                sector = row['Sector']
                if sector in sector_colors:
                    cell.fill = sector_colors[sector]
                
                if col_idx == df.columns.get_loc('Day Change %') + 1:
                    if value > 0:
                        cell.fill = self.colors['positive']
                    elif value < 0:
                        cell.fill = self.colors['negative']
        
        self._auto_adjust_columns(ws)
    
    def _create_technical_indicators_sheet(self, wb: Workbook, technical_indicators: List[TechnicalIndicators]):
        ws = wb.create_sheet("🔧 Technical Indicators")
        
        ws['A1'] = "TECHNICAL ANALYSIS INDICATORS"
        ws['A1'].font = self.fonts['title']
        
        if not technical_indicators:
            ws['A3'] = "No technical indicators available"
            return
        
        df = pd.DataFrame([{
            'Symbol': t.symbol,
            'RSI (14)': t.rsi,
            'SMA 20': t.sma_20,
            'SMA 50': t.sma_50,
            'SMA 200': t.sma_200,
            'Volatility (30d)': t.volatility,
            'Momentum 20d': t.momentum_20d,
            'Momentum 50d': t.momentum_50d,
            'Bollinger Upper': t.bollinger_upper,
            'Bollinger Lower': t.bollinger_lower,
            'MACD': t.macd,
            'Signal Strength': t.signal_strength
        } for t in technical_indicators])
        
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if col_idx == df.columns.get_loc('RSI (14)') + 1:
                    if value > 70:
                        cell.fill = self.colors['strong_down']
                    elif value < 30:
                        cell.fill = self.colors['strong_up']
                    else:
                        cell.fill = self.colors['neutral']
                
                elif col_idx == df.columns.get_loc('Signal Strength') + 1:
                    if 'STRONG_UP' in str(value):
                        cell.fill = self.colors['strong_up']
                    elif 'STRONG_DOWN' in str(value):
                        cell.fill = self.colors['strong_down']
                    elif 'UP' in str(value):
                        cell.fill = self.colors['positive']
                    elif 'DOWN' in str(value):
                        cell.fill = self.colors['negative']
        
        self._auto_adjust_columns(ws)
    
    def _create_sector_analysis_sheet(self, wb: Workbook, sector_analysis: List[SectorAnalysis]):
        ws = wb.create_sheet("🏭 Sector Analysis")
        
        ws['A1'] = "SECTOR PERFORMANCE ANALYSIS"
        ws['A1'].font = self.fonts['title']
        
        if not sector_analysis:
            ws['A3'] = "No sector analysis available"
            return
        
        df = pd.DataFrame([{
            'Sector': s.sector,
            'Stock Count': s.stock_count,
            'Avg Price Change %': s.avg_price_change,
            'Avg Volume Change %': s.avg_volume_change,
            'Sector Volatility': s.sector_volatility,
            'Trend Direction': s.trend_direction,
            'Sector Strength': s.sector_strength,
            'Top Performers': ', '.join(s.top_performers),
            'Bottom Performers': ', '.join(s.bottom_performers)
        } for s in sector_analysis])
        
        df = df.sort_values('Avg Price Change %', ascending=False)
        
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if col_idx == df.columns.get_loc('Avg Price Change %') + 1:
                    if value > 2:
                        cell.fill = self.colors['strong_up']
                    elif value > 0:
                        cell.fill = self.colors['positive']
                    elif value < -2:
                        cell.fill = self.colors['strong_down']
                    elif value < 0:
                        cell.fill = self.colors['negative']
                
                elif col_idx == df.columns.get_loc('Trend Direction') + 1:
                    if value == 'UPTREND':
                        cell.fill = self.colors['positive']
                    elif value == 'DOWNTREND':
                        cell.fill = self.colors['negative']
                    else:
                        cell.fill = self.colors['neutral']
        
        self._auto_adjust_columns(ws)
    
    def _create_historical_trends_sheet(self, wb: Workbook, stock_data: List[StockData], sector_analysis: List[SectorAnalysis]):
        ws = wb.create_sheet("📈 Historical Trends")
        
        ws['A1'] = "HISTORICAL TRENDS & PATTERNS"
        ws['A1'].font = self.fonts['title']
        
        ws['A3'] = "📊 Price Movement Summary"
        ws['A3'].font = self.fonts['subtitle']
        
        sectors = {}
        for stock in stock_data:
            if stock.sector not in sectors:
                sectors[stock.sector] = {'stocks': [], 'avg_change': 0}
            sectors[stock.sector]['stocks'].append(stock)
        
        summary_data = []
        for sector, data in sectors.items():
            stocks = data['stocks']
            avg_change = np.mean([stock.price_change_pct for stock in stocks])
            volatility = np.std([stock.price_change_pct for stock in stocks])
            
            summary_data.append({
                'Sector': sector,
                'Stock Count': len(stocks),
                'Average Change %': avg_change,
                'Price Volatility': volatility,
                'Best Performer': max(stocks, key=lambda x: x.price_change_pct).symbol,
                'Worst Performer': min(stocks, key=lambda x: x.price_change_pct).symbol
            })
        
        df = pd.DataFrame(summary_data)
        
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 6):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        ws['A15'] = "📝 TREND ANALYSIS NOTES"
        ws['A15'].font = self.fonts['subtitle']
        
        notes = [
            "• This analysis is based on current day's data",
            "• For comprehensive historical analysis, connect to historical database",
            "• Trends are identified using technical indicators and price movements",
            "• Sector rotation patterns can be observed from relative performance"
        ]
        
        for i, note in enumerate(notes):
            ws[f'A{17+i}'] = note
        
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


class MarketDataCollector:
    """Pure Market Data Collection - NO Customer Logic"""
    
    def __init__(self):
        self.sectors = {
            "TECH": [
                ("INFY.NS", "Infosys Limited"),
                ("TCS.NS", "Tata Consultancy Services"),
                ("WIPRO.NS", "Wipro Limited"),
                ("HCLTECH.NS", "HCL Technologies"),
                ("TECHM.NS", "Tech Mahindra"),
            ],
            "BANKING": [
                ("HDFCBANK.NS", "HDFC Bank"),
                ("ICICIBANK.NS", "ICICI Bank"),
                ("KOTAKBANK.NS", "Kotak Mahindra Bank"),
                ("AXISBANK.NS", "Axis Bank"),
                ("SBIN.NS", "State Bank of India"),
                ("INDUSINDBK.NS", "IndusInd Bank")
            ],
            "PHARMA": [
                ("SUNPHARMA.NS", "Sun Pharmaceutical"),
                ("DRREDDY.NS", "Dr. Reddy's Labs"),
                ("CIPLA.NS", "Cipla Limited"),
                ("DIVISLAB.NS", "Divi's Laboratories"),
                ("BIOCON.NS", "Biocon Limited"),
                ("LUPIN.NS", "Lupin Limited")
            ],
            "AUTO": [
                ("TATAMOTORS.NS", "Tata Motors"),
                ("MARUTI.NS", "Maruti Suzuki"),
                ("M&M.NS", "Mahindra & Mahindra"),
                ("BAJAJ-AUTO.NS", "Bajaj Auto"),
                ("EICHERMOT.NS", "Eicher Motors"),
                ("HEROMOTOCO.NS", "Hero MotoCorp")
            ],
            "ENERGY": [
                ("RELIANCE.NS", "Reliance Industries"),
                ("ONGC.NS", "Oil & Natural Gas Corp"),
                ("NTPC.NS", "NTPC Limited"),
                ("TATAPOWER.NS", "Tata Power"),
                ("ADANIGREEN.NS", "Adani Green Energy"),
                ("IOC.NS", "Indian Oil Corporation")
            ],
            "FMCG": [
                ("HINDUNILVR.NS", "Hindustan Unilever"),
                ("ITC.NS", "ITC Limited"),
                ("NESTLEIND.NS", "Nestle India"),
                ("BRITANNIA.NS", "Britannia Industries"),
                ("DABUR.NS", "Dabur India"),
                ("GODREJCP.NS", "Godrej Consumer Products")
            ]
        }
        
        # Build stock mappings
        self.all_stocks = []
        self.stock_names = {}
        self.stock_sectors = {}
        
        for sector, stocks in self.sectors.items():
            for symbol, name in stocks:
                self.all_stocks.append(symbol)
                self.stock_names[symbol] = name
                self.stock_sectors[symbol] = sector
        
        self.excel_reporter = ExcelMarketReporter()
        
        logger.info(f"Monitor Agent initialized: {len(self.all_stocks)} stocks, {len(self.sectors)} sectors")
    
    def collect_stock_data(self, symbols: List[str]) -> List[StockData]:
        """Collect stock data"""
        stock_data = []
        
        logger.info(f"Collecting data for {len(symbols)} stocks...")
        
        for symbol in symbols:
            try:
                stock = yf.Ticker(symbol)
                hist_data = stock.history(period="5d")
                
                if not hist_data.empty and len(hist_data) >= 2:
                    latest = hist_data.iloc[-1]
                    previous = hist_data.iloc[-2]
                    
                    price_change = latest['Close'] - previous['Close']
                    price_change_pct = (price_change / previous['Close']) * 100
                    
                    try:
                        info = stock.info
                        market_cap = info.get('marketCap', 0)
                        pe_ratio = info.get('forwardPE', 0)
                    except:
                        market_cap = 0
                        pe_ratio = 0
                    
                    stock_data_obj = StockData(
                        symbol=symbol,
                        company_name=self.stock_names.get(symbol, symbol.replace('.NS', '')),
                        sector=self.stock_sectors.get(symbol, 'UNKNOWN'),
                        current_price=float(latest['Close']),
                        price_change=float(price_change),
                        price_change_pct=float(price_change_pct),
                        volume=int(latest['Volume']),
                        market_cap=float(market_cap),
                        pe_ratio=float(pe_ratio) if pe_ratio else 0,
                        open_price=float(latest['Open']),
                        high_price=float(latest['High']),
                        low_price=float(latest['Low']),
                        date=latest.name.strftime('%Y-%m-%d')
                    )
                    
                    stock_data.append(stock_data_obj)
                    logger.info(f"✓ {symbol}")
                
            except Exception as e:
                logger.error(f"✗ {symbol}: {e}")
                continue
        
        logger.info(f"Collected {len(stock_data)} stocks")
        return stock_data
    
    def calculate_technical_indicators(self, symbols: List[str]) -> List[TechnicalIndicators]:
        """Calculate technical indicators"""
        technical_indicators = []
        
        logger.info(f"Calculating indicators for {len(symbols)} stocks...")
        
        for symbol in symbols:
            try:
                stock = yf.Ticker(symbol)
                hist_data = stock.history(period="1y")
                
                if len(hist_data) < 50:
                    continue
                
                prices = hist_data['Close'].values
                
                rsi = self._calculate_rsi(prices)
                sma_20 = np.mean(prices[-20:])
                sma_50 = np.mean(prices[-50:]) if len(prices) >= 50 else sma_20
                sma_200 = np.mean(prices[-200:]) if len(prices) >= 200 else sma_50
                
                volatility = self._calculate_volatility(prices)
                momentum_20d = self._calculate_momentum(prices, 20)
                momentum_50d = self._calculate_momentum(prices, 50)
                
                bollinger_upper, bollinger_lower = self._calculate_bollinger_bands(prices)
                macd = self._calculate_macd(prices)
                signal_strength = self._determine_signal_strength(prices, rsi, macd)
                
                tech_indicator = TechnicalIndicators(
                    symbol=symbol,
                    rsi=rsi,
                    sma_20=sma_20,
                    sma_50=sma_50,
                    sma_200=sma_200,
                    volatility=volatility,
                    momentum_20d=momentum_20d,
                    momentum_50d=momentum_50d,
                    bollinger_upper=bollinger_upper,
                    bollinger_lower=bollinger_lower,
                    macd=macd,
                    signal_strength=signal_strength
                )
                
                technical_indicators.append(tech_indicator)
                logger.info(f"✓ {symbol}")
                
            except Exception as e:
                logger.error(f"✗ {symbol}: {e}")
                continue
        
        logger.info(f"Calculated {len(technical_indicators)} indicators")
        return technical_indicators
    
    def analyze_sectors(self, stock_data: List[StockData]) -> List[SectorAnalysis]:
        """Analyze sector performance"""
        sector_groups = {}
        
        for stock in stock_data:
            if stock.sector not in sector_groups:
                sector_groups[stock.sector] = []
            sector_groups[stock.sector].append(stock)
        
        sector_analyses = []
        
        for sector, stocks in sector_groups.items():
            if not stocks:
                continue
            
            price_changes = [stock.price_change_pct for stock in stocks]
            
            avg_price_change = np.mean(price_changes)
            sector_volatility = np.std(price_changes)
            
            positive_count = len([p for p in price_changes if p > 0])
            negative_count = len([p for p in price_changes if p < 0])
            
            if positive_count > negative_count * 1.5:
                trend_direction = "UPTREND"
            elif negative_count > positive_count * 1.5:
                trend_direction = "DOWNTREND"
            else:
                trend_direction = "SIDEWAYS"
            
            sorted_stocks = sorted(stocks, key=lambda x: x.price_change_pct, reverse=True)
            top_performers = [stock.symbol for stock in sorted_stocks[:3]]
            bottom_performers = [stock.symbol for stock in sorted_stocks[-3:]]
            
            if avg_price_change > 2:
                sector_strength = "STRONG"
            elif avg_price_change > 0:
                sector_strength = "MODERATE"
            else:
                sector_strength = "WEAK"
            
            sector_analysis = SectorAnalysis(
                sector=sector,
                stock_count=len(stocks),
                avg_price_change=avg_price_change,
                avg_volume_change=0,
                sector_volatility=sector_volatility,
                trend_direction=trend_direction,
                top_performers=top_performers,
                bottom_performers=bottom_performers,
                sector_strength=sector_strength
            )
            
            sector_analyses.append(sector_analysis)
        
        logger.info(f"Analyzed {len(sector_analyses)} sectors")
        return sector_analyses
    
    def assess_market_overview(self, stock_data: List[StockData]) -> MarketOverview:
        """Assess overall market"""
        if not stock_data:
            return MarketOverview(
                analysis_date=datetime.now(),
                market_sentiment="NEUTRAL",
                market_volatility=0.0,
                advancing_stocks=0,
                declining_stocks=0,
                total_volume=0,
                market_breadth=1.0,
                fear_greed_index="NEUTRAL"
            )
        
        advancing_stocks = len([s for s in stock_data if s.price_change_pct > 0])
        declining_stocks = len([s for s in stock_data if s.price_change_pct < 0])
        total_volume = sum(s.volume for s in stock_data)
        
        market_breadth = advancing_stocks / declining_stocks if declining_stocks > 0 else 2.0
        
        price_changes = [s.price_change_pct for s in stock_data]
        market_volatility = np.std(price_changes) / 100
        
        avg_change = np.mean(price_changes)
        if avg_change > 1.0 and market_breadth > 1.2:
            market_sentiment = "BULLISH"
        elif avg_change < -1.0 and market_breadth < 0.8:
            market_sentiment = "BEARISH"
        else:
            market_sentiment = "NEUTRAL"
        
        if market_volatility > 0.03:
            fear_greed_index = "FEAR"
        elif market_volatility < 0.015 and avg_change > 0.5:
            fear_greed_index = "GREED"
        else:
            fear_greed_index = "NEUTRAL"
        
        return MarketOverview(
            analysis_date=datetime.now(),
            market_sentiment=market_sentiment,
            market_volatility=market_volatility,
            advancing_stocks=advancing_stocks,
            declining_stocks=declining_stocks,
            total_volume=total_volume,
            market_breadth=market_breadth,
            fear_greed_index=fear_greed_index
        )
    
    def _calculate_rsi(self, prices: np.ndarray, period: int = 14) -> float:
        """Calculate RSI"""
        if len(prices) < period + 1:
            return 50.0
        
        deltas = np.diff(prices)
        gains = np.where(deltas > 0, deltas, 0)
        losses = np.where(deltas < 0, -deltas, 0)
        
        avg_gain = np.mean(gains[-period:])
        avg_loss = np.mean(losses[-period:])
        
        if avg_loss == 0:
            return 100.0
        
        rs = avg_gain / avg_loss
        rsi = 100 - (100 / (1 + rs))
        return float(rsi)
    
    def _calculate_volatility(self, prices: np.ndarray, window: int = 30) -> float:
        """Calculate annualized volatility"""
        if len(prices) < 2:
            return 0.0
        
        returns = np.diff(np.log(prices))
        if len(returns) < window:
            volatility = np.std(returns) * np.sqrt(252)
        else:
            volatility = np.std(returns[-window:]) * np.sqrt(252)
        
        return float(volatility)
    
    def _calculate_momentum(self, prices: np.ndarray, period: int) -> float:
        """Calculate price momentum"""
        if len(prices) < period + 1:
            return 0.0
        
        momentum = (prices[-1] / prices[-period-1]) - 1
        return float(momentum)
    
    def _calculate_bollinger_bands(self, prices: np.ndarray, period: int = 20, std_dev: float = 2.0) -> tuple[float, float]:
        """Calculate Bollinger Bands"""
        if len(prices) < period:
            return float(prices[-1]), float(prices[-1])
        
        sma = np.mean(prices[-period:])
        std = np.std(prices[-period:])
        
        upper_band = sma + (std_dev * std)
        lower_band = sma - (std_dev * std)
        
        return float(upper_band), float(lower_band)
    
    def _calculate_macd(self, prices: np.ndarray, fast: int = 12, slow: int = 26) -> float:
        """Calculate MACD"""
        if len(prices) < slow:
            return 0.0
        
        ema_fast = self._calculate_ema(prices, fast)
        ema_slow = self._calculate_ema(prices, slow)
        
        macd = ema_fast - ema_slow
        return float(macd)
    
    def _calculate_ema(self, prices: np.ndarray, period: int) -> float:
        """Calculate Exponential Moving Average"""
        if len(prices) < period:
            return float(np.mean(prices))
        
        multiplier = 2.0 / (period + 1)
        ema = prices[0]
        
        for price in prices[1:]:
            ema = (price * multiplier) + (ema * (1 - multiplier))
        
        return float(ema)
    
    def _determine_signal_strength(self, prices: np.ndarray, rsi: float, macd: float) -> str:
        """Determine overall signal strength"""
        signals = []
        
        # RSI signals
        if rsi > 80:
            signals.append(-2)
        elif rsi > 70:
            signals.append(-1)
        elif rsi < 20:
            signals.append(2)
        elif rsi < 30:
            signals.append(1)
        else:
            signals.append(0)
        
        # MACD signals
        if macd > 0:
            signals.append(1)
        elif macd < 0:
            signals.append(-1)
        else:
            signals.append(0)
        
        # Price trend
        if len(prices) >= 5:
            recent_trend = (prices[-1] / prices[-5]) - 1
            if recent_trend > 0.02:
                signals.append(1)
            elif recent_trend < -0.02:
                signals.append(-1)
            else:
                signals.append(0)
        
        total_signal = sum(signals)
        
        if total_signal >= 3:
            return "STRONG_UP"
        elif total_signal >= 1:
            return "UP"
        elif total_signal <= -3:
            return "STRONG_DOWN"
        elif total_signal <= -1:
            return "DOWN"
        else:
            return "NEUTRAL"
    
    def run_comprehensive_data_collection(self) -> str:
        """Run complete market data collection - NO customer input"""
        logger.info("Starting comprehensive market data collection...")
        
        try:
            # Collect all market data
            all_stock_data = self.collect_stock_data(self.all_stocks)
            technical_indicators = self.calculate_technical_indicators(self.all_stocks)
            sector_analyses = self.analyze_sectors(all_stock_data)
            market_overview = self.assess_market_overview(all_stock_data)
            
            # Create general market report
            market_report = self.excel_reporter.create_market_report(
                market_overview=market_overview,
                stock_data=all_stock_data,
                technical_indicators=technical_indicators,
                sector_analysis=sector_analyses
            )
            
            # Export general market data for ALL agents to consume
            self._export_general_market_data(
                all_stock_data, technical_indicators, 
                sector_analyses, market_overview
            )
            
            logger.info("Market data collection completed successfully!")
            return market_report
            
        except Exception as e:
            logger.error(f"Error in data collection: {e}")
            raise
    
    def _export_general_market_data(self, stock_data: List[StockData], 
                                  technical_indicators: List[TechnicalIndicators],
                                  sector_analyses: List[SectorAnalysis],
                                  market_overview: MarketOverview):
        """Export general market data for all agents"""
        
        # Create general export directory
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
        export_dir = os.path.join(project_root, "data", "monitor_data")
        os.makedirs(export_dir, exist_ok=True)
        
        # Export general market data
        market_data = {
            "timestamp": datetime.now().isoformat(),
            "market_overview": {
                "sentiment": market_overview.market_sentiment,
                "volatility": market_overview.market_volatility,
                "breadth": market_overview.market_breadth,
                "fear_greed_index": market_overview.fear_greed_index
            },
            "stocks": [],
            "sectors": []
        }
        
        # Add all stock data
        for stock in stock_data:
            tech_data = next((t for t in technical_indicators if t.symbol == stock.symbol), None)
            
            stock_entry = {
                "symbol": stock.symbol,
                "company_name": stock.company_name,
                "sector": stock.sector,
                "current_price": stock.current_price,
                "price_change_pct": stock.price_change_pct,
                "volume": stock.volume,
                "market_cap": stock.market_cap,
                "technical_indicators": {
                    "rsi": tech_data.rsi if tech_data else None,
                    "volatility": tech_data.volatility if tech_data else None,
                    "momentum_20d": tech_data.momentum_20d if tech_data else None,
                    "signal_strength": tech_data.signal_strength if tech_data else None
                }
            }
            market_data["stocks"].append(stock_entry)
        
        # Add all sector data
        for sector in sector_analyses:
            sector_entry = {
                "sector": sector.sector,
                "performance": sector.avg_price_change,
                "volatility": sector.sector_volatility,
                "trend": sector.trend_direction,
                "strength": sector.sector_strength
            }
            market_data["sectors"].append(sector_entry)
        
        # Save general market data
        with open(f"{export_dir}/general_market_data.json", 'w') as f:
            json.dump(market_data, f, indent=2)
        
        logger.info(f"General market data exported to {export_dir}")


def main():
    """Main function to demonstrate customer-agnostic Monitor Agent"""
    logger.info("Starting Monitor Agent - Pure Market Data Collection")
    
    # Create data collector
    data_collector = MarketDataCollector()
    
    try:
        # Run comprehensive data collection (NO customer input)
        market_report = data_collector.run_comprehensive_data_collection()
        
        print("\n" + "="*80)
        print("🎉 MARKET DATA COLLECTION COMPLETE!")
        print("="*80)
        print(f"📊 Market Report: {market_report}")
        print("\n📋 Report Contains:")
        print("   📊 Market Overview - Market sentiment & conditions")
        print("   💰 All Stock Data - Complete price & volume data")
        print("   🔧 Technical Indicators - RSI, moving averages, signals")
        print("   🏭 Sector Analysis - All sector performance & trends")
        print("   📈 Historical Trends - Price patterns & analysis")
        print("\n📤 Data Export for Other Agents:")
        print("   📁 data/monitor_data/general_market_data.json")
        print("   📄 Complete market intelligence for Planner consumption")
        print("="*80)
        
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    main()